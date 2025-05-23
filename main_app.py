import os
import re
import io
import msoffcrypto
import pdfplumber
import numpy as np
import pandas as pd
from io import BytesIO
import streamlit as st
from openpyxl import load_workbook
from openpyxl.styles import Font
from datetime import datetime, timedelta




col1, col2, col3 = st.columns([1, 3, 1])
with col2:
    st.image("images/bci_logo.jpeg", width=250)

# Streamlit App Title
st.title("📄 Automation")

st.sidebar.image("images/itachi.jpg", width=250)  # Replace with your logo image path
st.sidebar.header("Navigation")
main_option = st.sidebar.radio("Choose an Option:", ["Prodigy", "Salary Finance"])



def is_valid_filename_warehouse(filename):
    pattern = r"^\d{10}-[A-Za-z]+-OnbehalfofitsCompartmentC\d{8}-\d{1,2}-[A-Za-z0-9]+\.pdf$"
    return bool(re.match(pattern, filename))

def is_valid_filename_abbsr(filename):
    pattern = r"^PFCM\s\d{4}-\d\sServicer\sReport\s-\s[A-Za-z]+\s\d{1,2}\.pdf$"
    return bool(re.match(pattern, filename))

def is_valid_filename_dfc(filename):
    pattern = r"^\d+\.\d+-\d+\sPFCM\s\d{4}-\d\sServicer\sReport\s-\s[A-Za-z]+\s\d{1,2}\s?\(.*\)?\.pdf$"
    return bool(re.match(pattern, filename))

if main_option == "Prodigy":
    option = st.selectbox(
    "Select the Report Type",
    ("Prodigy Warehouse", "Prodigy ABBSR", "Prodigy DFC")
    )
    if option == "Prodigy Warehouse":
        st.subheader("Upload Warehouse PDF")
        # File Uploader (PDF)
        uploaded_file = st.file_uploader("Upload a PDF file", type=["pdf"])

        if uploaded_file:
            filename = uploaded_file.name
            if is_valid_filename_warehouse(filename):
                st.success(f"✅ PDF '{filename}' uploaded successfully and matches format!")
                # Proceed with extraction process here
                # Create 'tables' folder if it doesn't exist
                folder_name = "tables"
                os.makedirs(folder_name, exist_ok=True)

                # Process the PDF file
                with pdfplumber.open(uploaded_file) as pdf:
                    formatted_date = "Unknown-Date"  # Default value
                    extracted_tables = []
                    
                    for page_num, page in enumerate(pdf.pages, start=1):
                        tables = page.extract_tables()
                        
                        if tables:
                            for table_idx, table in enumerate(tables, start=1):
                                df = pd.DataFrame(table)

                                # Extract the date from the first table on page 1
                                if page_num == 1 and table_idx == 1 and df.shape[1] > 1:
                                    raw_date = df.iloc[0, 1]
                                    try:
                                        date_obj = datetime.strptime(raw_date, "%d %B %Y")
                                        formatted_date = date_obj.strftime("%d-%b-%Y")
                                    except ValueError:
                                        formatted_date = "Unknown-Date"

                                # Save each table as an Excel file
                                output_filename = f"extracted_table_page_{page_num}_table_{table_idx}.xlsx"
                                file_path = os.path.join(folder_name, output_filename)
                                df.to_excel(file_path, index=False, header=False)
                                extracted_tables.append((file_path, df))

                # st.write(f"**Extracted Date:** {formatted_date}")

                # Compute adjusted dates for dynamic column names
                date_minus_1_month = (datetime.strptime(formatted_date, "%d-%b-%Y") - timedelta(days=29)).strftime("%d-%b-%Y")
                date_minus_2_months = (datetime.strptime(formatted_date, "%d-%b-%Y") - timedelta(days=91)).strftime("%d-%b-%Y")

                # Define column mappings for specific tables
                table_column_mappings = {
                "extracted_table_page_1_table_2.xlsx": [
                    "Bond", "ISIN", "Currency", "Notes Held", "Clean Price",
                    "Clean Price + Interest", "Remaining Principal", "Accrued Interest",
                    f"Balance {formatted_date}", "Pool factor"
                ],
                "extracted_table_page_1_table_3.xlsx": [
                    "Bond", "Currency", "Margin Above Base", f"Base Rate {date_minus_2_months}",
                    "Target Interest Rate", f"Balance {date_minus_1_month}",
                    "New Investment or Sale", "Interest Earned", "Interest Payment",
                    "Principal Payment", f"Balance {formatted_date}"
                ],
                "extracted_table_page_2_table_2.xlsx": [
                    "Bond", "ISIN", "Currency", "Notes Held", "Clean Price",
                    "Clean Price + Interest", "Remaining Principal", "Accrued Interest",
                    f"Balance {formatted_date}", "Pool factor"
                ],
                "extracted_table_page_2_table_3.xlsx": [
                    "Bond", "Currency", "Margin Above Base", f"Base Rate {date_minus_2_months}",
                    "Target Interest Rate", f"Balance {date_minus_1_month}",
                    "New Investment or Sale", "Interest Earned", "Interest Payment",
                    "Principal Payment", f"Balance {formatted_date}"
                ],
                "extracted_table_page_3_table_2.xlsx": [
                    "Bond", "ISIN", "Currency", "Notes Held", "Clean Price",
                    "Clean Price + Interest", "Remaining Principal", "Accrued Interest",
                    f"Balance {formatted_date}", "Pool factor"
                ],
                "extracted_table_page_4_table_2.xlsx": [
                    "Bond", "Currency", "Margin Above Base", f"Base Rate {date_minus_2_months}",
                    "Target Interest Rate", f"Balance {date_minus_1_month}",
                    "New Investment or Sale", "Interest Earned", "Interest Payment",
                    "Principal Payment", f"Balance {formatted_date}"
                ]
                }

                # Update column names based on mapping
                for file_name, columns in table_column_mappings.items():
                    try:
                        file_path = os.path.join(folder_name, file_name)
                        df = pd.read_excel(file_path, header=None)
                        df.columns = columns
                        df.to_excel(file_path, index=False)
                    except Exception as e:
                        st.error(f"Error updating {file_name}: {e}")

                # Merge all extracted tables into a single Excel file
                output_file = "prodigy_warehouse_output.xlsx"
                tables_folder = os.path.join(os.getcwd(), "tables")
                excel_files = sorted(
                    [f for f in os.listdir(tables_folder) if f.endswith('.xlsx') and 'table_1' not in f],
                    key=lambda x: (int(re.search(r'page_(\d+)', x).group(1)), int(re.search(r'table_(\d+)', x).group(1)))
                )

                with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
                    start_row = 0
                    for file in excel_files:
                        file_path = os.path.join(folder_name, file)
                        data = pd.read_excel(file_path)
                        data.to_excel(writer, sheet_name='Sheet1', startrow=start_row, index=False)
                        start_row += len(data) + 2  

                # Provide download button for final Excel file
                with open(output_file, "rb") as f:
                    st.download_button("📥 Download Processed Excel", f, file_name=f"{option.replace(' ', '_')}.xlsx", mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")

            else:
                st.error("❌ Invalid file format! Please upload a correctly formatted file.")

            


            # st.write("Processing Prodigy ABBSR Report...")
            # Call your Prodigy ABBSR function here
    elif option == "Prodigy ABBSR":
        st.subheader("Upload ABBSR PDF")
        # File Uploader (PDF)
        uploaded_file = st.file_uploader("Upload a PDF file", type=["pdf"])

        if uploaded_file:
            filename = uploaded_file.name
            if is_valid_filename_abbsr(filename):
                st.success(f"✅ PDF {filename}' uploaded successfully!")

                # Create 'tables' folder if it doesn't exist
                folder_name = "abbsr_files"
                os.makedirs(folder_name, exist_ok=True)    

                with pdfplumber.open(uploaded_file) as pdf:
                    page_2 = pdf.pages[1]  # Page index 1 for Page 2
                    text_2 = page_2.extract_text()  # Extract text instead of table
                    page_5 = pdf.pages[4]  # Page index 1 for Page 2
                    text_5 = page_5.extract_text()  # Extract text instead of table

                lines = text_2.split("\n")
                lines_page2 = text_5.split("\n")

                key_value_pairs = [
                    ("PRODIGY FINANCE CM 2021-1, DAC", "POOL SUMMARY"),
                    ("Distribution Date", lines[1].split()[2]),  # Extract the date
                    ("Collection Period End", lines[2].split()[3])  # Extract the date
                ]

                # Convert the key-value pairs into a DataFrame
                df1 = pd.DataFrame(key_value_pairs, columns=["Key", "Value"])
                output_filename = os.path.join(folder_name, 'extracted_df1.xlsx')
                df1.to_excel(output_filename, index=False, header=False)


                def fix_number_spacing(text):
                    text = text.replace(",", "")  # Remove all commas
                    
                    # Fix number spacing except for dates and percentages
                    text = re.sub(r'(\d{1,2}) (January|February|March|April|May|June|July|August|September|October|November|December) (\d{4})', r'\1_\2_\3', text)
                    text = re.sub(r'(\d) (\d{3,})\b', r'\1\2', text)  # Fix thousands/millions
                    text = re.sub(r'(\d) (\d\.\d+)', r'\1\2', text)  # Fix decimal numbers like 1 4.3 → 14.3
                    text = re.sub(r'(\d) \.(\d+)', r'\1.\2', text)  # Fix cases like '9 .9' → '9.9'
                    
                    return text

                def merge_split_lines(lines):
                    """Merge lines starting with 'Period' into the previous line."""
                    merged_lines = []
                    
                    for line in lines:
                        if line.strip().startswith("Period"):
                            # Merge with the last line in the list
                            if merged_lines:
                                merged_lines[-1] += " " + line.strip()  # Append it to the last line
                        else:
                            merged_lines.append(line.strip())  # Add new line as normal
                    
                    return merged_lines

                def merge_loan_balance_lines(lines):
                    """Merge lines starting with 'Total Modified Loan balance over Annual Period' with the next line if it starts with a number."""
                    merged_lines = []
                    i = 0

                    while i < len(lines):
                        line = lines[i].strip()

                        # Check if the current line contains the target phrase
                        if "Total Modified Loan balance over Annual Period" in line and i + 1 < len(lines):
                            next_line = lines[i + 1].strip()
                            
                            # Check if the next line starts with a number
                            if re.match(r'^\d', next_line):
                                line += " " + next_line  # Merge the next line with the current line
                                i += 1  # Skip the next line as it's merged

                        merged_lines.append(line)
                        i += 1

                    return merged_lines

                def handle_last_line(line):
                    """Fix the last line where 'Modified Month' is incorrectly spaced and numbers are joined."""
                    # Fix the spaced-out "Modified Month" issue
                    line = line.replace(",", "")
                    line = re.sub(r'\bMod i f i e d\b', 'Modified', line)
                    line = re.sub(r'\bM o n t h\b', 'Month', line)
                    line = line.strip()
                    parts = line.rsplit(" ", 2)  # Split by last 3 spaces
                    if len(parts) == 4 and parts[-2].isdigit() and parts[-1].isdigit():
                        first_part, num1, num2 = parts[0], parts[-2], parts[-1]
                        
                        return f"{first_part} {num1} {num2}"  # Format properly
                    
                    return line + " "

                merged_lines = merge_split_lines(lines)
                merged_lines = merge_loan_balance_lines(merged_lines)
                cleaned_lines = [fix_number_spacing(line) for line in merged_lines[4:-1]]

                if "Total Ever since" in lines[-1]:
                    lines[-1] = handle_last_line(lines[-1])
                    cleaned_lines.append(lines[-1])

                data = []
                for line in cleaned_lines:
                    parts = line.rsplit(" ", 3)  # Split into last 3 columns + description
                    if len(parts) == 4:
                        desc, num1, num2, num3 = parts
                        data.append([desc.strip(), num1.strip(), num2.strip(), num3.strip()])
                    else:
                        data.append([line, "", "", ""])  # Handle cases where format is different


                df2 = pd.DataFrame(data, columns=["Collateral Strats", "Current Reporting", "Previous Month", "7th June 2021 CutOff"])
                output_filename = os.path.join(folder_name, 'extracted_df2.xlsx')
                df2.to_excel(output_filename, index=False, header=True)

                # Page 5 cleaning
                pool_balance = lines_page2[5].replace(",", "").split()
                pool_balance = pool_balance[-2]+ pool_balance[-1]

                key_value_pairs = [
                    ("Delinquency & Payment Arrangement Data", " "),
                    ("Annualised Constant Prepayment Rate", lines_page2[1].split()[-1]),  # Extract the date
                    ("Annualised Constant Default Rate", lines_page2[2].split()[-1]),  # Extract the date
                    ("Cumulative Default rate", lines_page2[3].split()[-1]),
                    ("All balances are principal, admin fee and grace interest, unless stated otherwise", ""),
                    ("Pool Balance", pool_balance),
                    (lines_page2[6], lines_page2[7])
                ]

                # Convert the key-value pairs into a DataFrame
                df3 = pd.DataFrame(key_value_pairs, columns=["Key", "Value"])
                output_filename = os.path.join(folder_name, 'extracted_df3.xlsx')
                df3.to_excel(output_filename, index=False, header=False)


                def fix_number_spacing(text):
                    # Remove commas
                    text = text.replace(",", "")
                    def remove_spaces_from_last(text):

                        words = text.split()
                        if len(words) >= 3 and words[-2] != '-' and words[-3] != '-':
                            words[-3] = words[-3] + words[-2]  
                            words.pop(-2) 
                        
                        if len(words) >= 6 and words[-2] != '-' and words[-3] != '-':
                            words[-6] = words[-6] + words[-5]  
                            words.pop(-5)  
                        
                        return " ".join(words)
                    
                    return remove_spaces_from_last(text)



                cleaned_lines = [fix_number_spacing(line) for line in lines_page2[9:]]
                
                def remove_spaces_from_first(text):

                    words = text.split()
                    
                    if len(words) >= 6:
                        # Check if both words at -6 and -5 are numeric
                        if words[-6].replace('.', '', 1).isdigit() and words[-5].replace('.', '', 1).isdigit():
                            words[-6] = words[-6] + words[-5]  # Concatenate as a string
                            words.pop(-5)  # Remove the now unnecessary element

                    
                    return " ".join(words)

                cleaned_lines_final = [remove_spaces_from_first(line) for line in cleaned_lines]

                def process_numbers(text):
                    # Find all numbers in the text (including decimals and percentages)
                    words = text.split()  # Split text into words

                    for i in range(len(words)):
                        # Match numbers with more than 1 decimal place (excluding percentages)
                        if re.match(r'^\d+\.\d{2,}$', words[i]):  
                            original_number = words[i]  # Store the original number as a string
                            
                            rounded_number = "{:.1f}".format(float(original_number))  # Keep one decimal place as a string
                            
                            extra_value = str(int(round((float(original_number) - float(rounded_number)) * 100, 0)))  # Extract decimal part as string
                            
                            words[i] = rounded_number  # Replace with rounded number
                            
                            # Find the next valid number (excluding percentages) and add extra_value
                            for j in range(i+1, len(words)):
                                if re.match(r'^\d+(\.\d+)?$', words[j]):  # Ensure it's a valid number (not %)
                                    words[j] = str(extra_value) + words[j]
                                    break  # Stop after modifying the next number

                    return " ".join(words) 

                cleaned_lines_finall = [process_numbers(line) for line in cleaned_lines_final]

                columns = ["Delinquency - Total Pool", "Outstanding Principal", "# Loans", "% of Principal", "WA seasoning", "WA margin"]

                processed_data = []
                for row in cleaned_lines_finall[:9]:
                    words = row.split()  # Split by space
                    name = " ".join(words[:-5])  # Join everything except the last 5 numbers
                    values = words[-5:]  # Last 5 items are numeric values
                    processed_data.append([name] + values)

                # Create DataFrame
                df4 = pd.DataFrame(processed_data, columns=columns)
                output_filename = os.path.join(folder_name, 'extracted_df4.xlsx')
                df4.to_excel(output_filename, index=False, header=True)

                columns = ["Delinquency - Forbearance", "Outstanding Principal", "# Loans", "% of Principal", "WA seasoning", "WA margin"]

                processed_data = []
                for row in cleaned_lines_finall[10:]:
                    words = row.split()  # Split by space
                    name = " ".join(words[:-5])  # Join everything except the last 5 numbers
                    values = words[-5:]  # Last 5 items are numeric values
                    processed_data.append([name] + values)

                # Create DataFrame
                df5 = pd.DataFrame(processed_data, columns=columns)
                df5 = df5.drop(df5.index[-1]) 
                output_filename = os.path.join(folder_name, 'extracted_df5.xlsx')
                df5.to_excel(output_filename, index=False, header=True)

                output_file = 'prodigy_abbsr_pdf_output.xlsx'
                tables_folder = os.path.join(os.getcwd(), "abbsr_files")
                excel_files = [f for f in os.listdir(tables_folder) if f.endswith('.xlsx')]
                sorted_file_names = sorted(excel_files, key=lambda x: int(re.search(r'(\d+)', x).group()))
                # sorted_file_names
                with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
                    start_row = 0  # Track the starting row for each dataset
                    
                    for file in sorted_file_names:
                        # Read the data from the current file
                        file_path = os.path.join(folder_name, file)
                        data = pd.read_excel(file_path)
                        
                        # Write data to the final Excel file at the correct row position
                        data.to_excel(writer, sheet_name='Sheet1', startrow=start_row, index=False)

                        # Update start_row to place the next dataset after 2 empty rows
                        start_row += len(data) + 2 

                # Provide download button for final Excel file
                with open(output_file, "rb") as f:
                    st.download_button("📥 Download Processed Excel", f, file_name=f"{option.replace(' ', '_')}.xlsx", mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")

            else:
                st.error("❌ Invalid file format! Please upload a correctly formatted file.")


    elif option == "Prodigy DFC":
        st.subheader("Upload DFC PDF")
        # File Uploader (PDF)
        uploaded_file = st.file_uploader("Upload a PDF file", type=["pdf"])
        if uploaded_file:
            filename = uploaded_file.name
            if is_valid_filename_dfc(filename):
                st.success(f"✅ PDF {filename}' uploaded successfully!")

                # Create 'tables' folder if it doesn't exist
                folder_name = "dfc_files"
                os.makedirs(folder_name, exist_ok=True)

                with pdfplumber.open(uploaded_file) as pdf:
                    page_1 = pdf.pages[0]  # Page index 1 for Page 2
                    text_1 = page_1.extract_text()  # Extract text instead of table
                    page_5 = pdf.pages[4]
                    text_5 = page_5.extract_text()
                    page_6 = pdf.pages[5]
                    text_6 = page_6.extract_text()

                lines = text_1.split("\n")
                lines_page5 = text_5.split("\n")
                lines_page6 = text_6.split("\n")

                key_value_pairs = [
                    (lines[1], ""),
                    (lines[2], "")
                ]

                # Convert the key-value pairs into a DataFrame
                df1 = pd.DataFrame(key_value_pairs, columns=["Key", "Value"])
                output_filename = os.path.join(folder_name, 'extracted_df1.xlsx')
                df1.to_excel(output_filename, index=False, header=False)

                def is_number(s):
                    """Check if a string represents a number (integer or float)."""
                    try:
                        float(s)  # Try converting to float
                        return True
                    except ValueError:
                        return False

                def process_strings_limited(s):
                    s = s.replace(",", "")  # Remove all commas
                    words = s.split()

                    # Check if the last two words are numbers (integer or float)
                    if len(words) >= 2 and is_number(words[-1]) and is_number(words[-2]):
                        words[-2] = words[-2] + words[-1]  # Concatenate numbers
                        words.pop()  # Remove the last word
                    
                    return " ".join(words)

                cleaned_lines = [process_strings_limited(line) for line in lines[4:26]]

                data = []
                for line in cleaned_lines:
                    words = line.split()
                    key = " ".join(words[:-1])  # Everything except last word
                    value = words[-1]  # Last word is the value
                    data.append((key, value))

                df2 = pd.DataFrame(data, columns=[lines[3], ""])
                # df2
                output_filename = os.path.join(folder_name, 'extracted_df2.xlsx')
                df2.to_excel(output_filename, index=False, header=False)

                def clean_number_format(line):
                    # Remove commas
                    line = line.replace(",", "")
                    
                    # Split the line into words
                    parts = line.split()
                    
                    # Ensure there are at least three elements to check
                    if len(parts) > 3:
                        # Check if the last three parts are numbers (including decimals)
                        if re.match(r'^\d+(\.\d+)?$', parts[-1]) and re.match(r'^\d+(\.\d+)?$', parts[-2]) and re.match(r'^\d+(\.\d+)?$', parts[-3]):
                            # Concatenate -3 and -2
                            parts[-3] = parts[-3] + parts[-2]
                            parts.pop(-2)  # Remove the now redundant -2
                    
                    return " ".join(parts)

                cleaned_lines = [clean_number_format(line) for line in lines[26:-1]]

                data = []
                for line in cleaned_lines:
                    parts = line.split()
                    if len(parts) >= 3 and parts[-1].isdigit() and parts[-2].isdigit():
                        name = " ".join(parts[:-2])  # Everything except last two parts
                        num1 = parts[-2]  # Second last part
                        num2 = parts[-1]  # Last part
                        data.append([name, num1, num2])
                    else:
                        data.append([line, "", ""])  # Keep the first column, leave the others empty

                # Create DataFrame
                df3 = pd.DataFrame(data, columns=["", "", ""])
                # df3
                output_filename = os.path.join(folder_name, 'extracted_df3.xlsx')
                df3.to_excel(output_filename, index=False, header=False)


                def clean_number_format(line):
                    # Remove commas
                    line = line.replace(",", "")

                    # Split the line into words
                    parts = line.split()

                    # Iterate through the words and merge numbers separated by spaces
                    cleaned_parts = []
                    i = 0
                    while i < len(parts):
                        # Check if current and next part are numbers
                        if i < len(parts) - 1 and re.match(r'^\d+$', parts[i]) and re.match(r'^\d+$', parts[i+1]):
                            cleaned_parts.append(parts[i] + parts[i+1])  # Merge numbers
                            i += 2  # Skip next part as it's merged
                        else:
                            cleaned_parts.append(parts[i])
                            i += 1

                    return " ".join(cleaned_parts)

                # Example usage
                cleaned_lines = [clean_number_format(line) for line in lines_page5[26:-1]]

                def split_key_value(line):
                    parts = line.rsplit(" ", 1)  # Split at the last space
                    last_part = parts[-1].strip()

                    # Check if last part is a number (including negative numbers in parentheses) or PASS/FAIL
                    if re.match(r'^-?\d+$', last_part) or re.match(r'^\(-?\d+\)$', last_part) or re.match(r'^\d{2}/\d{2}/\d{4}$', last_part) or last_part in {"PASS", "FAIL", "-"}:
                        key = parts[0] if len(parts) > 1 else ""
                        value = last_part
                    else:
                        key = line  # Whole line as key
                        value = ""

                    return key, value

                data = [split_key_value(line) for line in cleaned_lines]
                df4 = pd.DataFrame(data, columns=["", ""])
                # df4
                output_filename = os.path.join(folder_name, 'extracted_df4.xlsx')
                df4.to_excel(output_filename, index=False, header=False)

                def clean_number_format(line):
                    # Remove commas
                    line = line.replace(",", "")

                    # Split the line into words
                    parts = line.split()

                    # Iterate through the words and merge numbers separated by spaces
                    cleaned_parts = []
                    i = 0
                    while i < len(parts):
                        # Check if current and next part are numbers
                        if i < len(parts) - 1 and re.match(r'^\d+$', parts[i]) and re.match(r'^\d+$', parts[i+1]):
                            cleaned_parts.append(parts[i] + parts[i+1])  # Merge numbers
                            i += 2  # Skip next part as it's merged
                        else:
                            cleaned_parts.append(parts[i])
                            i += 1

                    return " ".join(cleaned_parts)

                # Example usage
                cleaned_lines = [clean_number_format(line) for line in lines_page6[1:34]]

                def split_key_value(line):
                    parts = line.rsplit(" ", 1)  # Split at the last space
                    last_part = parts[-1].strip()

                    # Check if last part is a number (including negative numbers in parentheses) or PASS/FAIL
                    if  re.match(r'^-?\d+$', last_part) or re.match(r'^\(-?\d+\)$', last_part) or re.match(r'^\d{2}/\d{2}/\d{4}$', last_part) or re.match(r'^-?\d+(\.\d+)?%$', last_part) or re.match(r'^-?\d+\.\d+$', last_part) or last_part in {"PASS", "FAIL", "-"}:
                        key = parts[0] if len(parts) > 1 else ""
                        value = last_part
                    else:
                        key = line  # Whole line as key
                        value = ""

                    return key, value

                data = [split_key_value(line) for line in cleaned_lines]
                df5 = pd.DataFrame(data, columns=["", ""])
                # df5
                output_filename = os.path.join(folder_name, 'extracted_df5.xlsx')
                df5.to_excel(output_filename, index=False, header=False)

                output_file = 'prodigy_dfc_pdf_output.xlsx'
                tables_folder = os.path.join(os.getcwd(), "dfc_files")
                excel_files = [f for f in os.listdir(tables_folder) if f.endswith('.xlsx')]
                sorted_file_names = sorted(excel_files, key=lambda x: int(re.search(r'(\d+)', x).group()))
                # sorted_file_names
                with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
                    start_row = 0  # Track the starting row for each dataset
                    
                    for file in sorted_file_names:
                        # Read the data from the current file
                        file_path = os.path.join(folder_name, file)
                        # print("file_path", file_path)
                        data = pd.read_excel(file_path)
                        
                        # Write data to the final Excel file at the correct row position
                        data.to_excel(writer, sheet_name='Sheet1', startrow=start_row, index=False)

                        # Update start_row to place the next dataset after 2 empty rows
                        start_row += len(data) + 2

                # Provide download button for final Excel file
                with open(output_file, "rb") as f:
                    st.download_button("📥 Download Processed Excel", f, file_name=f"{option.replace(' ', '_')}.xlsx", mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")

            else:
                st.error("❌ Invalid file format! Please upload a correctly formatted file.")

if main_option == "Salary Finance":
    base_file = st.file_uploader("Upload a Base Report file", type=["xlsb"])
    servicer_file = st.file_uploader("Upload a Servicer Report file", type=["xlsb"])
    
    if base_file and servicer_file:
        base_filename = base_file.name
        servicer_filename = servicer_file.name

        if "Skylark" in base_filename and "Base" in base_filename and "Skylark" in servicer_filename and "Servicer" in servicer_filename:
            st.success("Both files are uploaded and validated successfully!")

            with st.spinner("Processing Data...Please wait."):

                def get_password(servicer_filename):
                    """
                    Extract password and file date from the filename.
                    Assumes filename starts with date in 'YYYY.MM.DD' format.
                    """
                    # filename = os.path.basename(file_path)
                    match = re.match(r'(\d{4}\.\d{2}\.\d{2})', servicer_filename)
                    if match:
                        date_part = match.group(1)
                        date_without_dots = date_part.replace('.', '')
                        password_servicer = "SRSkylark" + date_without_dots
                        password_base = "BBRSkylark" + date_without_dots
                        # st.write(f"Password for Servicer Report: {password_servicer}")
                        # st.write(f"Password for Base Report: {password_base}")
                        # filedate = date_part  # Return date as string
                        return password_servicer, password_base
                    else:
                        raise ValueError(f"Date not found at the beginning of filename: {servicer_filename}")

                # Read Base Report file
                sheets_base = [
                    "11. Sub BB Schedule", "7. Concentration limits", "5. Advance Rate", "4. Mezz BB Schedule", 
                    "3. Snr BB Schedule", "0. Pre-Funding Forecast", "8.B. Employer concentration", "1.B. Data Tape"
                ]
                password_servicer, password_base = get_password(servicer_filename)  # Set the actual password
                decrypted_file_base = BytesIO()
                office_file = msoffcrypto.OfficeFile(base_file)
                # st.write(office_file)
                office_file.load_key(password=password_base)
                office_file.decrypt(decrypted_file_base)

                dfs_base = {sheet: pd.read_excel(decrypted_file_base, sheet_name=sheet, engine="pyxlsb") for sheet in sheets_base}
                
                # Read Servicer Report file (Decrypted)
                sheets_servicer = [
                    "13. Seller Related Events", "12. Asset Related Events", "5. Advance Rate", "3. Snr BB Schedule"
                ]
                # password = get_password(servicer_filename)  # Set the actual password
                decrypted_file_servicer = BytesIO()
                office_file = msoffcrypto.OfficeFile(servicer_file)
                office_file.load_key(password=password_servicer)
                office_file.decrypt(decrypted_file_servicer)

                dfs_servicer = {sheet: pd.read_excel(decrypted_file_servicer, sheet_name=sheet, engine="pyxlsb") for sheet in sheets_servicer}

                # Extract DataFrames
                df_sub_bb = dfs_base["11. Sub BB Schedule"]
                df_conc_lim = dfs_base["7. Concentration limits"]
                df_adv_rate = dfs_base["5. Advance Rate"]
                df_mezz_bb = dfs_base["4. Mezz BB Schedule"]
                df_sen_bb = dfs_base["3. Snr BB Schedule"]
                df_pre_fund = dfs_base["0. Pre-Funding Forecast"]
                df_emp_conc = dfs_base["8.B. Employer concentration"]
                df_data_tape = dfs_base["1.B. Data Tape"]

                df_seller = dfs_servicer["13. Seller Related Events"]
                df_asset = dfs_servicer["12. Asset Related Events"]
                df_advance = dfs_servicer["5. Advance Rate"]
                df_senior = dfs_servicer["3. Snr BB Schedule"]

                # Calculations
                bb1 = df_mezz_bb.iloc[15, 5] + df_mezz_bb.iloc[20, 5] 
                bb2 = df_mezz_bb.iloc[23:26, 5].sum()
                bb3 = df_mezz_bb.iloc[26:28, 5].sum()
                bb4 = df_sen_bb.iloc[37, 5] + df_sen_bb.iloc[35, 5]
                bb5 = df_mezz_bb.iloc[34, 5] + df_mezz_bb.iloc[32, 5]
                bb6 = bb1 + bb2 + bb3 - (bb4 + bb5)

                bb7 = df_pre_fund.iloc[47, 4]
                bb8 = df_sub_bb.iloc[22, 5]
                bb9 = bb7 + bb8
                bb10 = df_sub_bb.iloc[27, 5]
                bb11 = bb10 - bb9

                def format_percent(value):
                    if pd.isna(value) or value == "-" or value == "":
                        return ""  
                    return f"{float(value) * 100:.2f}%"

                cc1 = df_conc_lim.iloc[12, 5]
                cc2 = format_percent(df_conc_lim.iloc[14, 5])
                cc3 = format_percent(df_emp_conc.iloc[8, 7]) if df_emp_conc.iloc[16, 7] == df_conc_lim.iloc[17, 2] else "N/A"
                cc4 = format_percent(df_emp_conc.iloc[8, 7]) if df_emp_conc.iloc[16, 7] == df_conc_lim.iloc[18, 2] else "N/A"
                cc5 = format_percent(df_emp_conc.iloc[13, 7])

                numerator_6 = df_data_tape.loc[(df_data_tape["MethodOfPayment"].isin(["Push Payment", "Direct Debit"])) & (df_data_tape["IS ineglible"] == 0), "PrincipalBalanceOutstanding"].sum()
                denominator_6 = df_data_tape.loc[df_data_tape["IS ineglible"] == 0, "PrincipalBalanceOutstanding"].sum()
                cc6 = format_percent(numerator_6 / denominator_6 if denominator_6 != 0 else 0)

                numerator_7 = np.where((df_data_tape["MethodOfPayment"] == "Push Payment") & (df_data_tape["IS ineglible"] == 0), df_data_tape["PrincipalBalanceOutstanding"], 0).sum()
                denominator_7 = np.where(df_data_tape["IS ineglible"] == 0, df_data_tape["PrincipalBalanceOutstanding"], 0).sum()
                cc7 = format_percent(numerator_7 / denominator_7 if denominator_7 != 0 else 0)

                numerator_8 = df_data_tape.loc[(df_data_tape["IsPositiveCounterOffer"] == 1) & (df_data_tape["DelphiCreditBucket"] == "Prime") & (df_data_tape["IS ineglible"] == 0), "PrincipalBalanceOutstanding"].sum()
                denominator_8 = df_data_tape.loc[df_data_tape["IS ineglible"] == 0, "PrincipalBalanceOutstanding"].sum()
                cc8 = format_percent(numerator_8 / denominator_8 if denominator_8 != 0 else 0)

                numerator_9 = df_data_tape.loc[(df_data_tape["PCOwIncreasedMnthlyPmt"] == 1) & (df_data_tape["DelphiCreditBucket"] == "Near Prime") & (df_data_tape["IS ineglible"] == 0), "PrincipalBalanceOutstanding"].sum()
                denominator_9 = df_data_tape.loc[df_data_tape["IS ineglible"] == 0, "PrincipalBalanceOutstanding"].sum()
                cc9 = format_percent(numerator_9 / denominator_9 if denominator_9 != 0 else 0)

                numerator_10 = (df_data_tape["IS ineglible"].eq(0) & df_data_tape["DelphiCreditBucket"].eq("Subprime")) * df_data_tape["PrincipalBalanceOutstanding"]
                numerator_sum = numerator_10.sum()
                denominator_10 = df_data_tape.loc[df_data_tape["IS ineglible"] == 0, "PrincipalBalanceOutstanding"].sum()
                cc10 = format_percent(numerator_sum / denominator_10 if denominator_10 != 0 else 0)

                df_row_27 = pd.to_numeric(df_seller.loc[27], errors="coerce").dropna()
                df_row_27 = pd.to_datetime(df_row_27, origin='1899-12-30', unit='D')

                last_date_column = df_row_27.last_valid_index()

                last_valid_col_number = df_seller.columns.get_loc(last_date_column)

                fc1 = df_seller.iloc[52, last_valid_col_number]
                fc2 = df_seller.iloc[64, last_valid_col_number]

                fc3 = format_percent(df_advance.iloc[132, 4:7].mean())
                fc4 = format_percent(df_advance.iloc[122, 4:7].mean())
                fc5 = format_percent(df_advance.iloc[82, 5])
                fc6 = df_asset.iloc[15, 7]

                sc1 = format_percent(df_adv_rate.iloc[86, 5])
                sc2 = df_senior.iloc[37, 5]

                data = {
                    "Metric": [
                        "Raw Data",
                        "",
                        "Compliance Certificate", "Quarter", "Covenants", "Concentration Covenants",
                        "Average Weighted Remaing Term",
                        "Minimum WA APR",
                        "Single Employer (if <100m)",
                        "Single Employer (if >100m)",
                        "Top 5 Employer partenr",
                        "Direct Debit/Push Payments",
                        "Push Payments only",
                        "Adjusted Loan Terms Prime Borrowers only",
                        "Adjusted Loan Terms Near Prime Borrowers only",
                        "Sub-Prime Borrowers",
                        "",
                        "Financial/Performance Covenants",
                        "Tangible net worth",
                        "Liquidity and Min Cash",
                        "Deliquency Rate",
                        "Default Rate",
                        "Payment Not By Salary Deduction",
                        "Balance of Interest Reserve Account",
                        "Senior Covenants",
                        "Default Rate",
                        "Senior Drawn ",
                        "",
                        "Borrowing Base",
                        "",
                        "Subordinate Borrowing Base", "",
                        "the Eligible Receivables Balance as at the immediately preceding Cut-Off Date; plus ",
                        "the sum of:",
                        "the aggregate sum of the balance standing to the credit of all of the Issuer Accounts",
                        "the aggregate sum of the balance standing to the credit of all of the Sandpit Accounts and Disbursement Accounts",
                        "the principal amount outstanding of the Senior Loan on such date; and ",
                        "the principal amount outstanding of the Mezzanine Note on such date",
                        "Subordinated Borrowing Base = (a) + (b) - (c)",
                        "Subordinate funding request for valuation on reporting date",
                        "Current balance outstanding on subordinate funding line (before current funding request)",
                        "",
                        "Total commitment drawn to date (inlcuding current advance request)",
                        "",
                        "Total agreed facility commitment ",
                        "",
                        "Total funds available to be drawn (after current advance request)"
                    ],
                    "Value": [
                        "",
                        "",
                        "",  "", "", "",
                        cc1, cc2, cc3, cc4, cc5, cc6, cc7, cc8, cc9, cc10,
                        "",
                        "",
                        fc1, fc2, fc3, fc4, fc5, fc6,
                        "",
                        sc1, sc2,
                        "",
                        "", "", "", "",
                        bb1, "", bb2, bb3, bb4, bb5, bb6, bb7, bb8, 
                        "", bb9, "", bb10, "", bb11
                    ]
    }
                df_final = pd.DataFrame(data)

                # Save to Excel
                output_file = "salaryfinance_monitoring_sheet.xlsx"
                df_final.to_excel(output_file, index=False, header=False)

                # Apply formatting
                wb = load_workbook(output_file)
                ws = wb.active
                for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=2):
                    metric_cell, value_cell = row
                    if value_cell.value in ["", None]:  
                        metric_cell.font = Font(bold=True)

                wb.save(output_file)

            # Provide download button for final Excel file
            with open(output_file, "rb") as f:
                st.download_button("📥 Download Processed Excel", f, file_name="salaryfinance_monitoring_sheet.xlsx", mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")

        else:
            st.error("Invalid file names! Please upload the correct Base and Servicer report files.")


